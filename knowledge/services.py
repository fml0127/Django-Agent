import hashlib
import json
import math
import re
import tempfile
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path

import httpx
from bs4 import BeautifulSoup
from django.conf import settings
from django.db import connection, transaction
from openai import OpenAI
from sqlite_vec import serialize_float32

from content_runtime.converters import (
    ArchiveExtractionError,
    DocumentConversionError,
    convert_office_to_pdf,
    convert_legacy_office_to_pdf,
    extract_archive_members,
)
from content_runtime.extractors import VisionExtractionError, extract_image_file, extract_pdf_file_as_images
from content_runtime.inspectors import (
    HTML_SUFFIXES,
    RTF_SUFFIXES,
    TEXT_SUFFIXES,
    XLSX_SUFFIXES,
    FileProfile,
    inspect_bytes,
    inspect_stored_file,
)

from .models import ContentExtraction, KBChunk, KBDocument
from .sqlite_search import ensure_search_tables, vector_dim


class UnsupportedDocumentError(Exception):
    def __init__(self, message, failure_code="unsupported_format", profile=None, parser_name="", metadata=None):
        super().__init__(message)
        self.failure_code = failure_code
        self.profile = profile
        self.parser_name = parser_name
        self.metadata = metadata or {}


class DocumentParseError(Exception):
    def __init__(self, message, failure_code="parser_exception", profile=None, parser_name="", metadata=None):
        super().__init__(message)
        self.failure_code = failure_code
        self.profile = profile
        self.parser_name = parser_name
        self.metadata = metadata or {}


@dataclass
class Entry:
    raw: str
    content: str
    compiled: str
    title: str
    source: str
    metadata: dict = field(default_factory=dict)


@dataclass
class SearchHit:
    score: float
    chunk: KBChunk
    query: str = ""
    source_scores: dict = field(default_factory=dict)
    rerank_score: float | None = None

    def __iter__(self):
        yield self.score
        yield self.chunk


@dataclass
class ParsedFileResult:
    title: str
    entries: list[Entry]
    profile: FileProfile
    method: str
    parser_name: str
    normalized_text: str
    raw_output: str = ""
    model_name: str = ""
    failure_code: str = ""
    metadata: dict = field(default_factory=dict)


@dataclass
class ExtractedEntriesResult:
    entries: list[Entry]
    parser_name: str
    method: str
    metadata: dict = field(default_factory=dict)
    model_name: str = ""


def short_hash(text):
    return hashlib.sha1((text or "").encode("utf-8")).hexdigest()[:16]


def clean_text(text):
    text = (text or "").replace("\x00", "")
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line).strip()


def compact_text(text):
    return re.sub(r"\s+", " ", text or "").strip()


def _json_safe(value):
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, dict):
        return {str(key): _json_safe(val) for key, val in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    return str(value)


def normalize_metadata(metadata):
    return {str(key): _json_safe(value) for key, value in (metadata or {}).items()}


def _setting_int(name, default):
    try:
        return int(getattr(settings, name, default))
    except (TypeError, ValueError):
        return default


def _text_quality(entries, fallback_reason=""):
    entry_count = len(entries or [])
    total_chars = sum(len(clean_text(getattr(entry, "content", "") or "")) for entry in entries or [])
    avg_entry_chars = total_chars / entry_count if entry_count else 0
    min_total = _setting_int("OFFICE_MIN_TOTAL_TEXT_CHARS", 200)
    min_avg = _setting_int("OFFICE_MIN_AVG_ENTRY_TEXT_CHARS", 30)
    sparse = total_chars < min_total or avg_entry_chars < min_avg
    return {
        "total_chars": total_chars,
        "entry_count": entry_count,
        "avg_entry_chars": round(avg_entry_chars, 2),
        "min_total_chars": min_total,
        "min_avg_entry_chars": min_avg,
        "sparse": bool(sparse),
        "fallback_reason": fallback_reason or ("sparse_text" if sparse else ""),
    }


def _apply_text_quality(entries, quality):
    for entry in entries or []:
        entry.metadata["text_quality"] = normalize_metadata(quality)


def split_text(text, chunk_size=800, overlap=120):
    text = clean_text(text)
    if not text:
        return []
    from langchain_text_splitters import RecursiveCharacterTextSplitter

    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=overlap,
        separators=["\n\n", "\n", "。", "！", "？", ".", "!", "?", " ", ""],
        length_function=len,
    )
    return [chunk for chunk in (clean_text(part) for part in splitter.split_text(text)) if chunk]


def fallback_embedding(text, dim=96):
    values = [0.0] * dim
    for token in re.findall(r"[\w\u4e00-\u9fff]+", (text or "").lower()):
        idx = int(hashlib.md5(token.encode("utf-8")).hexdigest(), 16) % dim
        values[idx] += 1.0
    return normalize_vector(values)


def target_dim():
    return vector_dim()


def fit_vector(vector):
    dim = target_dim()
    if len(vector) == dim:
        return vector
    if len(vector) > dim:
        return vector[:dim]
    return vector + [0.0] * (dim - len(vector))


def normalize_vector(vector):
    values = [float(v) for v in fit_vector(vector)]
    norm = math.sqrt(sum(v * v for v in values))
    if not norm:
        return values
    return [v / norm for v in values]


def embed_text(text):
    if not settings.EMBEDDING_API_KEY:
        return fallback_embedding(text, dim=target_dim())
    try:
        client = OpenAI(api_key=settings.EMBEDDING_API_KEY, base_url=settings.EMBEDDING_BASE_URL)
        resp = client.embeddings.create(model=settings.EMBEDDING_MODEL, input=text or "")
        return normalize_vector(resp.data[0].embedding)
    except Exception:
        return fallback_embedding(text, dim=target_dim())


def _documents_to_entries(documents, title, source):
    entries = []
    for index, document in enumerate(documents):
        content = clean_text(getattr(document, "page_content", ""))
        if not content:
            continue
        metadata = normalize_metadata(getattr(document, "metadata", {}) or {})
        entry_title = metadata.get("title") or metadata.get("source") or title or source
        compiled = f"# {entry_title}\n{content}"
        metadata.update({"entry_index": index, "title": entry_title, "source": source})
        entries.append(
            Entry(
                raw=content,
                content=content,
                compiled=compiled,
                title=str(entry_title),
                source=source,
                metadata=metadata,
            )
        )
    if not entries:
        raise DocumentParseError("未解析出可入库文本。", failure_code="empty_text")
    return entries


def entries_from_text(text, title, source):
    content = clean_text(text)
    if not content:
        raise DocumentParseError("未解析出可入库文本。", failure_code="empty_text")
    return [
        Entry(
            raw=content,
            content=content,
            compiled=f"# {title}\n{content}",
            title=title,
            source=source,
            metadata={"title": title, "source": source},
        )
    ]


def parse_url_entries(url):
    try:
        from langchain_community.document_loaders import WebBaseLoader

        documents = WebBaseLoader(url).load()
    except Exception as exc:
        raise DocumentParseError(f"URL 解析失败：{exc}") from exc

    title = url
    for document in documents:
        metadata = getattr(document, "metadata", {}) or {}
        if metadata.get("title"):
            title = metadata["title"]
            break
    return title, _documents_to_entries(documents, title, url)


def parse_url(url):
    title, entries = parse_url_entries(url)
    return title, "\n\n".join(entry.content for entry in entries)


def _decode_text_bytes(data):
    encodings = ["utf-8", "utf-8-sig", "gb18030"]
    try:
        import chardet

        detected = (chardet.detect(data or b"") or {}).get("encoding")
        if detected:
            encodings.insert(0, detected)
    except Exception:
        pass
    encodings.append("latin-1")
    last_error = None
    for encoding in dict.fromkeys(encodings):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError as exc:
            last_error = exc
    raise DocumentParseError(f"文本编码识别失败：{last_error}", failure_code="encoding_failed")


def _entries_from_loader_documents(documents, title, source, profile, parser_name, method):
    entries = _documents_to_entries(documents, title, source)
    for entry in entries:
        entry.metadata.update(
            {
                "extraction_method": method,
                "parser_name": parser_name,
                "detected_mime": profile.mime,
                "detected_family": profile.family,
            }
        )
    return entries


def _loader_for_profile(profile, file_path):
    if profile.family == "text" or profile.suffix in TEXT_SUFFIXES:
        from langchain_community.document_loaders import TextLoader

        return TextLoader(file_path, encoding="utf-8", autodetect_encoding=True), "TextLoader"
    if profile.family == "html" or profile.suffix in HTML_SUFFIXES:
        from langchain_community.document_loaders import BSHTMLLoader

        return BSHTMLLoader(file_path), "BSHTMLLoader"
    if profile.family == "pdf":
        from langchain_community.document_loaders import PyMuPDFLoader

        return PyMuPDFLoader(file_path), "PyMuPDFLoader"
    if profile.family == "docx":
        from langchain_community.document_loaders import Docx2txtLoader

        return Docx2txtLoader(file_path), "Docx2txtLoader"
    if profile.family == "pptx":
        from langchain_community.document_loaders import UnstructuredPowerPointLoader

        return UnstructuredPowerPointLoader(file_path, mode="elements"), "UnstructuredPowerPointLoader"
    raise UnsupportedDocumentError(
        profile.reason or f"暂不支持 {profile.suffix or '无扩展名'} 格式。",
        failure_code=profile.failure_code or "unsupported_format",
        profile=profile,
    )


def _fallback_text_entries(file_bytes, title, source, profile):
    try:
        text, encoding = _decode_text_bytes(file_bytes)
    except DocumentParseError as exc:
        exc.profile = profile
        exc.parser_name = "TextEncodingFallback"
        raise
    try:
        entries = entries_from_text(text, title, source)
    except DocumentParseError as exc:
        exc.profile = profile
        exc.parser_name = "TextEncodingFallback"
        raise
    for entry in entries:
        entry.metadata.update(
            {
                "extraction_method": ContentExtraction.METHOD_TEXT_FALLBACK,
                "parser_name": "TextEncodingFallback",
                "encoding": encoding,
                "detected_mime": profile.mime,
                "detected_family": profile.family,
            }
        )
    return entries, "TextEncodingFallback", ContentExtraction.METHOD_TEXT_FALLBACK, {"encoding": encoding}


def _fallback_html_entries(file_bytes, title, source, profile):
    text, encoding = _decode_text_bytes(file_bytes)
    soup = BeautifulSoup(text, "html.parser")
    content = soup.get_text("\n")
    try:
        entries = entries_from_text(content, title, source)
    except DocumentParseError as exc:
        exc.profile = profile
        exc.parser_name = "BeautifulSoupFallback"
        raise
    for entry in entries:
        entry.metadata.update(
            {
                "extraction_method": ContentExtraction.METHOD_HTML_FALLBACK,
                "parser_name": "BeautifulSoupFallback",
                "encoding": encoding,
                "detected_mime": profile.mime,
                "detected_family": profile.family,
            }
        )
    return entries, "BeautifulSoupFallback", ContentExtraction.METHOD_HTML_FALLBACK, {"encoding": encoding}


def _simple_rtf_to_text(text):
    text = re.sub(r"\\'[0-9a-fA-F]{2}", " ", text or "")
    text = re.sub(r"\\[a-zA-Z]+-?\d* ?", " ", text)
    text = text.replace("\\{", "{").replace("\\}", "}").replace("\\\\", "\\")
    text = re.sub(r"[{}]", " ", text)
    return clean_text(text)


def _rtf_entries(file_bytes, title, source, profile):
    parser_name = "StripRtfExtractor"
    try:
        from striprtf.striprtf import rtf_to_text
    except Exception as exc:
        rtf_to_text = None
        parser_name = "SimpleRtfFallback"
        import_error = str(exc)
    else:
        import_error = ""

    text, encoding = _decode_text_bytes(file_bytes)
    try:
        content = rtf_to_text(text) if rtf_to_text else _simple_rtf_to_text(text)
    except Exception as exc:
        raise DocumentParseError(
            f"RTF 解析失败：{exc}",
            failure_code="parser_exception",
            profile=profile,
            parser_name=parser_name,
        ) from exc
    try:
        entries = entries_from_text(content, title, source)
    except DocumentParseError as exc:
        exc.profile = profile
        exc.parser_name = parser_name
        raise
    for entry in entries:
        entry.metadata.update(
            {
                "extraction_method": ContentExtraction.METHOD_TEXT_FALLBACK,
                "parser_name": parser_name,
                "encoding": encoding,
                "detected_mime": profile.mime,
                "detected_family": profile.family,
            }
        )
    return ExtractedEntriesResult(
        entries=entries,
        parser_name=parser_name,
        method=ContentExtraction.METHOD_TEXT_FALLBACK,
        metadata={"encoding": encoding, "striprtf_import_error": import_error},
    )


def _format_spreadsheet_cell(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()


def _sheet_rows_to_text(title, sheet_name, rows):
    lines = [f"# {title}", f"Sheet: {sheet_name}", ""]
    lines.extend("\t".join(row) for row in rows)
    return clean_text("\n".join(lines))


def _xlsx_entries_with_openpyxl(file_bytes, title, source, profile):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise DocumentParseError(
            f"缺少 openpyxl，无法解析 XLSX：{exc}",
            failure_code="spreadsheet_parse_failed",
            profile=profile,
            parser_name="SpreadsheetExtractor",
        ) from exc

    try:
        workbook = load_workbook(BytesIO(file_bytes or b""), read_only=True, data_only=True)
    except Exception as exc:
        raise DocumentParseError(
            f"openpyxl 解析 XLSX 失败：{exc}",
            failure_code="parser_exception",
            profile=profile,
            parser_name="SpreadsheetExtractor",
        ) from exc

    entries = []
    workbook_metadata = []
    try:
        for index, sheet in enumerate(workbook.worksheets):
            rows = []
            non_empty_cells = 0
            max_columns = 0
            for raw_row in sheet.iter_rows(values_only=True):
                formatted = [_format_spreadsheet_cell(cell) for cell in raw_row]
                while formatted and formatted[-1] == "":
                    formatted.pop()
                if not any(formatted):
                    continue
                non_empty_cells += sum(1 for cell in formatted if cell)
                max_columns = max(max_columns, len(formatted))
                rows.append(formatted)
            if not rows or non_empty_cells <= 0:
                continue
            normalized_rows = [row + [""] * (max_columns - len(row)) for row in rows]
            content = _sheet_rows_to_text(title, sheet.title, normalized_rows)
            if not content:
                continue
            cell_range = getattr(sheet, "calculate_dimension", lambda: "")() or ""
            sheet_metadata = {
                "sheet_name": sheet.title,
                "row_count": len(rows),
                "column_count": max_columns,
                "non_empty_cells": non_empty_cells,
                "cell_range": cell_range,
                "entry_index": index,
                "title": f"{title} - {sheet.title}",
                "source": source,
                "extraction_method": ContentExtraction.METHOD_TEXT_FALLBACK,
                "parser_name": "SpreadsheetExtractor",
                "detected_mime": profile.mime,
                "detected_family": profile.family,
            }
            workbook_metadata.append(
                {
                    "sheet_name": sheet.title,
                    "row_count": len(rows),
                    "column_count": max_columns,
                    "non_empty_cells": non_empty_cells,
                    "cell_range": cell_range,
                }
            )
            entries.append(
                Entry(
                    raw=content,
                    content=content,
                    compiled=content,
                    title=f"{title} - {sheet.title}",
                    source=source,
                    metadata=sheet_metadata,
                )
            )
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    if not entries:
        raise DocumentParseError(
            "XLSX 工作簿没有可入库的有效单元格。",
            failure_code="empty_text",
            profile=profile,
            parser_name="SpreadsheetExtractor",
        )
    return ExtractedEntriesResult(
        entries=entries,
        parser_name="SpreadsheetExtractor",
        method=ContentExtraction.METHOD_TEXT_FALLBACK,
        metadata={"sheets": workbook_metadata, "sheet_count": len(workbook_metadata)},
    )


def _xlsx_entries_with_langchain(file_bytes, title, source, profile):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
        tmp.write(file_bytes or b"")
        tmp.flush()
        try:
            from langchain_community.document_loaders.excel import UnstructuredExcelLoader
        except Exception:
            from langchain_community.document_loaders import UnstructuredExcelLoader

        loader = UnstructuredExcelLoader(tmp.name, mode="elements")
        documents = loader.load()
    entries = _entries_from_loader_documents(
        documents,
        title,
        source,
        profile,
        "UnstructuredExcelLoader",
        ContentExtraction.METHOD_LANGCHAIN,
    )
    return ExtractedEntriesResult(
        entries=entries,
        parser_name="UnstructuredExcelLoader",
        method=ContentExtraction.METHOD_LANGCHAIN,
    )


def _xlsx_entries(file_bytes, title, source, profile):
    try:
        return _xlsx_entries_with_openpyxl(file_bytes, title, source, profile)
    except DocumentParseError as openpyxl_exc:
        if openpyxl_exc.failure_code == "empty_text":
            raise
        try:
            extracted = _xlsx_entries_with_langchain(file_bytes, title, source, profile)
        except Exception as langchain_exc:
            raise DocumentParseError(
                f"表格解析失败：openpyxl={openpyxl_exc}; langchain={langchain_exc}",
                failure_code="spreadsheet_parse_failed",
                profile=profile,
                parser_name="SpreadsheetExtractor",
                metadata={
                    "openpyxl_error": str(openpyxl_exc)[:500],
                    "langchain_error": str(langchain_exc)[:500],
                },
            ) from langchain_exc
        extracted.metadata = {
            **(extracted.metadata or {}),
            "openpyxl_error": str(openpyxl_exc)[:500],
            "fallback": "UnstructuredExcelLoader",
        }
        return extracted


def _vision_entries_from_text(text, title, source, profile, parser_name):
    try:
        entries = entries_from_text(text, title, source)
    except DocumentParseError as exc:
        exc.profile = profile
        exc.parser_name = parser_name
        raise
    for entry in entries:
        entry.metadata.update(
            {
                "extraction_method": ContentExtraction.METHOD_VISION,
                "parser_name": parser_name,
                "model": settings.VISION_MODEL,
                "detected_mime": profile.mime,
                "detected_family": profile.family,
            }
        )
    return entries


def _extract_with_vision(tmp_path, title, source, profile):
    try:
        if profile.family == "image":
            text = extract_image_file(tmp_path, profile.mime, source)
            parser_name = "VisionImageExtractor"
        elif profile.family == "pdf":
            text = extract_pdf_file_as_images(tmp_path, source)
            parser_name = "VisionPdfImageExtractor"
        else:
            raise VisionExtractionError("当前文件类型不能使用视觉模型解析。", "vision_not_supported")
    except VisionExtractionError as exc:
        raise DocumentParseError(str(exc), failure_code=exc.failure_code, profile=profile, parser_name="VisionExtractor") from exc
    except Exception as exc:
        raise DocumentParseError(f"视觉模型解析失败：{exc}", failure_code="vision_parse_failed", profile=profile, parser_name="VisionExtractor") from exc

    entries = _vision_entries_from_text(text, title, source, profile, parser_name)
    return ExtractedEntriesResult(
        entries=entries,
        parser_name=parser_name,
        method=ContentExtraction.METHOD_VISION,
        metadata={"model": settings.VISION_MODEL},
        model_name=settings.VISION_MODEL,
    )


def _temp_suffix(profile, source):
    suffix = profile.suffix or Path(source or "").suffix.lower()
    return suffix if suffix else ".bin"


def _loader_entries_from_bytes(file_bytes, title, source, profile):
    with tempfile.NamedTemporaryFile(suffix=_temp_suffix(profile, source), delete=True) as tmp:
        tmp.write(file_bytes)
        tmp.flush()
        if profile.family == "image":
            return _extract_with_vision(tmp.name, title, source, profile)
        try:
            loader, parser_name = _loader_for_profile(profile, tmp.name)
            documents = loader.load()
            entries = _entries_from_loader_documents(
                documents,
                title,
                source,
                profile,
                parser_name,
                ContentExtraction.METHOD_LANGCHAIN,
            )
            return ExtractedEntriesResult(
                entries=entries,
                parser_name=parser_name,
                method=ContentExtraction.METHOD_LANGCHAIN,
            )
        except DocumentParseError as exc:
            if profile.family == "pdf" and exc.failure_code in {"parser_exception", "empty_text"}:
                return _extract_with_vision(tmp.name, title, source, profile)
            exc.profile = exc.profile or profile
            raise
        except Exception as exc:
            if profile.family == "pdf":
                return _extract_with_vision(tmp.name, title, source, profile)
            if profile.family == "text":
                entries, parser_name, method, parser_metadata = _fallback_text_entries(file_bytes, title, source, profile)
                return ExtractedEntriesResult(entries, parser_name, method, parser_metadata)
            if profile.family == "html":
                entries, parser_name, method, parser_metadata = _fallback_html_entries(file_bytes, title, source, profile)
                return ExtractedEntriesResult(entries, parser_name, method, parser_metadata)
            raise DocumentParseError(
                f"解析失败：{exc}",
                failure_code="parser_exception",
                profile=profile,
                parser_name=parser_name if "parser_name" in locals() else "",
            ) from exc


def _has_non_empty_entries(result):
    return bool(result and clean_text("\n\n".join(entry.content for entry in result.entries)))


def _extract_converted_pdf_entries(converted, title, source, original_profile, fallback_reason="", force_sparse_vision=False):
    converted_profile = inspect_bytes(f"{Path(source or title).stem or 'converted'}.pdf", "application/pdf", converted.data[:8192])
    try:
        extracted = _loader_entries_from_bytes(converted.data, title, source, converted_profile)
    except DocumentParseError:
        raise

    quality = _text_quality(extracted.entries, fallback_reason=fallback_reason)
    _apply_text_quality(extracted.entries, quality)
    if force_sparse_vision and extracted.method != ContentExtraction.METHOD_VISION and quality["sparse"]:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=True) as tmp:
            tmp.write(converted.data or b"")
            tmp.flush()
            try:
                vision_extracted = _extract_with_vision(tmp.name, title, source, converted_profile)
            except DocumentParseError as exc:
                if quality["total_chars"] > 0:
                    extracted.metadata = {
                        **(extracted.metadata or {}),
                        "text_quality": quality,
                        "vision_fallback": {
                            "attempted": True,
                            "status": "failed",
                            "failure_code": exc.failure_code,
                            "message": str(exc)[:500],
                        },
                    }
                    return extracted, converted_profile
                raise
        vision_quality = _text_quality(vision_extracted.entries, fallback_reason=fallback_reason)
        _apply_text_quality(vision_extracted.entries, vision_quality)
        vision_extracted.metadata = {
            **(vision_extracted.metadata or {}),
            "text_quality": vision_quality,
            "pdf_text_quality": quality,
            "vision_fallback": {"attempted": True, "status": "success"},
        }
        return vision_extracted, converted_profile

    extracted.metadata = {**(extracted.metadata or {}), "text_quality": quality}
    return extracted, converted_profile


def _apply_conversion_metadata(extracted, converted, converted_profile, original_profile, original_text_quality=None):
    conversion_metadata = converted.as_metadata()
    for entry in extracted.entries:
        entry.metadata.update(
            {
                "original_detected_mime": original_profile.mime,
                "original_detected_family": original_profile.family,
                "original_suffix": original_profile.suffix,
                "conversion": conversion_metadata,
            }
        )
        if original_text_quality:
            entry.metadata["original_text_quality"] = normalize_metadata(original_text_quality)
    extracted.parser_name = f"LibreOfficeConverter+{extracted.parser_name}"[:128]
    extracted.metadata = {
        **(extracted.metadata or {}),
        "conversion": conversion_metadata,
        "converted_profile": converted_profile.as_metadata(),
    }
    if original_text_quality:
        extracted.metadata["original_text_quality"] = original_text_quality
    return extracted


def _conversion_entries_from_bytes(file_bytes, title, source, profile):
    try:
        converted = convert_legacy_office_to_pdf(file_bytes, source)
    except DocumentConversionError as exc:
        if exc.failure_code == "needs_conversion":
            raise UnsupportedDocumentError(
                str(exc),
                failure_code=exc.failure_code,
                profile=profile,
                parser_name="LibreOfficeConverter",
                metadata=exc.metadata,
            ) from exc
        raise DocumentParseError(
            str(exc),
            failure_code=exc.failure_code,
            profile=profile,
            parser_name="LibreOfficeConverter",
            metadata=exc.metadata,
        ) from exc

    extracted, converted_profile = _extract_converted_pdf_entries(converted, title, source, profile)
    return _apply_conversion_metadata(extracted, converted, converted_profile, profile)


def _sparse_office_conversion_entries(file_bytes, title, source, profile, original_text_quality=None):
    try:
        converted = convert_office_to_pdf(file_bytes, source, source_family=profile.family)
    except DocumentConversionError as exc:
        if exc.failure_code == "needs_conversion":
            raise UnsupportedDocumentError(
                str(exc),
                failure_code=exc.failure_code,
                profile=profile,
                parser_name="LibreOfficeConverter",
                metadata={
                    **(exc.metadata or {}),
                    "fallback_reason": "sparse_text",
                    "original_text_quality": original_text_quality or {},
                },
            ) from exc
        raise DocumentParseError(
            f"稀疏文本兜底失败：{exc}",
            failure_code="sparse_text_fallback_failed",
            profile=profile,
            parser_name="LibreOfficeConverter",
            metadata={
                **(exc.metadata or {}),
                "conversion_failure_code": exc.failure_code,
                "fallback_reason": "sparse_text",
                "original_text_quality": original_text_quality or {},
            },
        ) from exc

    extracted, converted_profile = _extract_converted_pdf_entries(
        converted,
        title,
        source,
        profile,
        fallback_reason="sparse_text",
        force_sparse_vision=True,
    )
    extracted.metadata = {
        **(extracted.metadata or {}),
        "sparse_fallback": {"attempted": True, "status": "success", "target_format": "pdf"},
    }
    return _apply_conversion_metadata(extracted, converted, converted_profile, profile, original_text_quality)


def _office_loader_with_sparse_fallback(file_bytes, title, source, profile):
    original = None
    original_error = None
    try:
        original = _loader_entries_from_bytes(file_bytes, title, source, profile)
    except DocumentParseError as exc:
        original_error = exc

    if original is not None:
        quality = _text_quality(original.entries)
        _apply_text_quality(original.entries, quality)
        original.metadata = {**(original.metadata or {}), "text_quality": quality}
        if not quality["sparse"]:
            return original
        try:
            return _sparse_office_conversion_entries(file_bytes, title, source, profile, original_text_quality=quality)
        except (UnsupportedDocumentError, DocumentParseError) as fallback_exc:
            if quality["total_chars"] > 0:
                fallback_info = {
                    "attempted": True,
                    "status": "failed",
                    "failure_code": fallback_exc.failure_code,
                    "message": str(fallback_exc)[:500],
                }
                for entry in original.entries:
                    entry.metadata["sparse_fallback"] = fallback_info
                original.metadata = {
                    **(original.metadata or {}),
                    "sparse_fallback": fallback_info,
                }
                return original
            raise

    try:
        return _sparse_office_conversion_entries(file_bytes, title, source, profile)
    except (UnsupportedDocumentError, DocumentParseError):
        if original_error is not None and original_error.failure_code not in {"empty_text", "parser_exception"}:
            raise original_error
        raise


def _archive_entries_from_bytes(file_bytes, title, source, profile):
    try:
        members = extract_archive_members(file_bytes, source)
    except ArchiveExtractionError as exc:
        if exc.failure_code == "archive_unsupported":
            raise UnsupportedDocumentError(
                str(exc),
                failure_code=exc.failure_code,
                profile=profile,
                parser_name="ArchiveExtractor",
                metadata=exc.metadata,
            ) from exc
        raise DocumentParseError(
            str(exc),
            failure_code=exc.failure_code,
            profile=profile,
            parser_name="ArchiveExtractor",
            metadata=exc.metadata,
        ) from exc

    entries = []
    parser_names = []
    methods = []
    member_results = []
    member_failures = []
    for index, member in enumerate(members):
        member_profile = inspect_bytes(member.name, "", member.data[:8192])
        if member_profile.family == "archive":
            member_failures.append(
                {
                    "name": member.name,
                    "failure_code": "archive_unsupported",
                    "message": "不递归解析嵌套压缩包。",
                }
            )
            continue
        try:
            extracted = _extract_entries_from_bytes(member.data, Path(member.name).name, member.name, member_profile)
        except (UnsupportedDocumentError, DocumentParseError) as exc:
            member_failures.append(
                {
                    "name": member.name,
                    "failure_code": exc.failure_code,
                    "message": str(exc)[:500],
                }
            )
            continue
        parser_names.append(extracted.parser_name)
        methods.append(extracted.method)
        member_results.append(
            {
                "name": member.name,
                "family": member_profile.family,
                "parser": extracted.parser_name,
                "entry_count": len(extracted.entries),
                "metadata": member.metadata,
            }
        )
        for entry in extracted.entries:
            entry.source = f"{source}::{member.name}"
            entry.metadata.update(
                {
                    "archive_source": source,
                    "archive_member": member.name,
                    "archive_member_index": index,
                    "archive_member_profile": member_profile.as_metadata(),
                    **(member.metadata or {}),
                }
            )
            entries.append(entry)

    if not entries:
        raise DocumentParseError(
            "压缩包中没有可入库文件。",
            failure_code="archive_no_supported_files",
            profile=profile,
            parser_name="ArchiveExtractor",
        )
    parser_name = "ArchiveExtractor+" + "+".join(dict.fromkeys(parser_names or ["mixed"]))
    method = next((method for method in methods if method == ContentExtraction.METHOD_LANGCHAIN), methods[0] if methods else ContentExtraction.METHOD_TEXT_FALLBACK)
    return ExtractedEntriesResult(
        entries=entries,
        parser_name=parser_name[:128],
        method=method,
        metadata={
            "archive": {
                "member_count": len(members),
                "parsed_member_count": len(member_results),
                "members": member_results[:50],
                "failures": member_failures[:50],
            }
        },
    )


def _extract_entries_from_bytes(file_bytes, title, source, profile):
    if profile.parser_mode == "unsupported":
        raise UnsupportedDocumentError(
            profile.reason or f"暂不支持 {profile.suffix or '无扩展名'} 格式。",
            failure_code=profile.failure_code or "unsupported_format",
            profile=profile,
        )
    if profile.family == "rtf" or profile.suffix in RTF_SUFFIXES:
        return _rtf_entries(file_bytes, title, source, profile)
    if profile.family == "xlsx" or profile.suffix in XLSX_SUFFIXES:
        return _xlsx_entries(file_bytes, title, source, profile)
    if profile.family == "legacy_office":
        return _conversion_entries_from_bytes(file_bytes, title, source, profile)
    if profile.family in {"docx", "pptx"}:
        return _office_loader_with_sparse_fallback(file_bytes, title, source, profile)
    if profile.family == "archive":
        return _archive_entries_from_bytes(file_bytes, title, source, profile)
    return _loader_entries_from_bytes(file_bytes, title, source, profile)


def parse_user_file_result(user_file):
    stored = user_file.stored_file
    if not stored or not stored.file:
        raise DocumentParseError("文件内容不存在。")

    profile = inspect_stored_file(stored, save=True)
    try:
        with stored.file.open("rb") as stored_file:
            file_bytes = stored_file.read()
    except Exception as exc:
        raise DocumentParseError(f"读取文件失败：{exc}", failure_code="read_failed", profile=profile) from exc

    title = user_file.name
    source = user_file.name
    extracted = _extract_entries_from_bytes(file_bytes, title, source, profile)

    for entry in extracted.entries:
        entry.metadata.update(
            {
                "user_file_id": user_file.id,
                "filename": user_file.name,
                "suffix": profile.suffix,
                "mime_type": user_file.mime_type,
            }
        )
    normalized_text = clean_text("\n\n".join(entry.content for entry in extracted.entries))
    if not normalized_text:
        raise DocumentParseError(
            "未解析出可入库文本。",
            failure_code="empty_text",
            profile=profile,
            parser_name=extracted.parser_name,
        )
    return ParsedFileResult(
        title=title,
        entries=extracted.entries,
        profile=profile,
        method=extracted.method,
        parser_name=extracted.parser_name,
        normalized_text=normalized_text,
        raw_output="\n\n".join(entry.raw for entry in extracted.entries),
        model_name=extracted.model_name,
        metadata={
            "profile": profile.as_metadata(),
            "parser": extracted.parser_name,
            **(extracted.metadata or {}),
        },
    )


def parse_user_file_entries(user_file):
    result = parse_user_file_result(user_file)
    return result.title, result.entries


def parse_user_file(user_file):
    title, entries = parse_user_file_entries(user_file)
    return title, "\n\n".join(entry.content for entry in entries)


def quote_fts_query(query):
    escaped = compact_text(query).replace('"', '""')
    return f'"{escaped}"'


def delete_chunk_indexes(chunk_id):
    ensure_search_tables()
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM knowledge_kbchunk_vec WHERE chunk_id = %s", [chunk_id])
        cursor.execute("DELETE FROM knowledge_kbchunk_fts WHERE rowid = %s", [chunk_id])


def upsert_chunk_indexes(chunk, vector):
    ensure_search_tables()
    serialized_vector = serialize_float32(normalize_vector(vector))
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM knowledge_kbchunk_vec WHERE chunk_id = %s", [chunk.id])
        cursor.execute(
            "INSERT INTO knowledge_kbchunk_vec(chunk_id, kb_id, embedding) VALUES (%s, %s, %s)",
            [chunk.id, chunk.kb_id, serialized_vector],
        )
        cursor.execute("DELETE FROM knowledge_kbchunk_fts WHERE rowid = %s", [chunk.id])
        cursor.execute(
            "INSERT INTO knowledge_kbchunk_fts(rowid, content) VALUES (%s, %s)",
            [chunk.id, chunk.content],
        )


def vector_candidates(kb, query_vector, limit):
    ensure_search_tables()
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT chunk_id, distance
            FROM knowledge_kbchunk_vec
            WHERE embedding MATCH %s AND kb_id = %s AND k = %s
            ORDER BY distance
            """,
            [serialize_float32(normalize_vector(query_vector)), kb.id, max(1, int(limit))],
        )
        return [(int(chunk_id), float(distance)) for chunk_id, distance in cursor.fetchall()]


def fts_candidates(kb, query, limit):
    if len(compact_text(query)) < 3:
        return []
    ensure_search_tables()
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT knowledge_kbchunk_fts.rowid, bm25(knowledge_kbchunk_fts) AS rank
            FROM knowledge_kbchunk_fts
            JOIN knowledge_kbchunk AS chunk ON chunk.id = knowledge_kbchunk_fts.rowid
            WHERE knowledge_kbchunk_fts MATCH %s AND chunk.kb_id = %s
            ORDER BY rank
            LIMIT %s
            """,
            [quote_fts_query(query), kb.id, max(1, int(limit))],
        )
        return [(int(chunk_id), float(rank)) for chunk_id, rank in cursor.fetchall()]


def refresh_kb_doc_count(kb):
    kb.doc_count = kb.documents.filter(status=KBDocument.STATUS_READY).count()
    kb.save(update_fields=["doc_count", "updated_at"])


def delete_existing_file_documents(kb, user_file):
    existing = KBDocument.objects.filter(kb=kb, user_file=user_file)
    extraction_ids = [item for item in existing.values_list("extraction_id", flat=True) if item]
    existing.delete()
    ContentExtraction.objects.filter(kb=kb, user_file=user_file).delete()
    if extraction_ids:
        ContentExtraction.objects.filter(id__in=extraction_ids).delete()
    refresh_kb_doc_count(kb)


def chunk_entries(entries, chunk_size=800, overlap=120):
    chunk_payloads = []
    for entry_index, entry in enumerate(entries):
        for chunk_index, content in enumerate(split_text(entry.compiled, chunk_size=chunk_size, overlap=overlap)):
            metadata = dict(entry.metadata)
            metadata.update(
                {
                    "entry_index": entry_index,
                    "entry_chunk_index": chunk_index,
                    "title": entry.title,
                    "source": entry.source,
                }
            )
            chunk_payloads.append((content, metadata))
    return chunk_payloads


def mark_document_failed(doc, status, message, failure_code="index_failed"):
    KBChunk.objects.filter(document=doc).delete()
    doc.status = status
    doc.error_message = str(message)[:2000]
    doc.chunk_count = 0
    doc.failure_code = failure_code
    doc.save(update_fields=["status", "error_message", "chunk_count", "failure_code", "updated_at"])
    refresh_kb_doc_count(doc.kb)
    return doc


def create_content_extraction(
    kb,
    user_file,
    status,
    method,
    normalized_text="",
    raw_output="",
    failure_code="",
    error_message="",
    model_name="",
    metadata=None,
):
    return ContentExtraction.objects.create(
        kb=kb,
        user_file=user_file,
        status=status,
        method=method,
        model_name=(model_name or "")[:128],
        raw_output=raw_output or "",
        normalized_text=normalized_text or "",
        failure_code=(failure_code or "")[:64],
        metadata=metadata or {},
        error_message=str(error_message or "")[:2000],
    )


def create_status_document(
    kb,
    source_type,
    source,
    title,
    status,
    message,
    user_file=None,
    content_hash="",
    extraction=None,
    profile=None,
    parser_name="",
    failure_code="",
    parser_metadata=None,
    replace_existing=True,
):
    if replace_existing and source_type == KBDocument.SOURCE_FILE and user_file:
        delete_existing_file_documents(kb, user_file)
    doc = KBDocument.objects.create(
        kb=kb,
        source_type=source_type,
        source=source,
        user_file=user_file,
        extraction=extraction,
        title=(title or source)[:512],
        content_hash=(content_hash or "")[:64],
        status=status,
        error_message=str(message)[:2000],
        chunk_count=0,
        detected_mime=(profile.mime if profile else "")[:128],
        detected_family=(profile.family if profile else "")[:32],
        parser_name=(parser_name or "")[:128],
        failure_code=(failure_code or "")[:64],
        parser_metadata=parser_metadata or {},
    )
    refresh_kb_doc_count(kb)
    return doc


def ingest_entries(
    kb,
    source_type,
    source,
    title,
    entries,
    user_file=None,
    content_hash="",
    chunk_size=800,
    overlap=120,
    extraction=None,
    profile=None,
    parser_name="",
    parser_metadata=None,
    replace_existing=True,
):
    ensure_search_tables()
    if replace_existing and source_type == KBDocument.SOURCE_FILE and user_file:
        delete_existing_file_documents(kb, user_file)

    doc = KBDocument.objects.create(
        kb=kb,
        source_type=source_type,
        source=source,
        user_file=user_file,
        extraction=extraction,
        title=(title or source)[:512],
        content_hash=(content_hash or short_hash("\n".join(entry.compiled for entry in entries)))[:64],
        status=KBDocument.STATUS_PROCESSING,
        detected_mime=(profile.mime if profile else "")[:128],
        detected_family=(profile.family if profile else "")[:32],
        parser_name=(parser_name or "")[:128],
        parser_metadata=parser_metadata or {},
    )
    chunk_payloads = chunk_entries(entries, chunk_size=chunk_size, overlap=overlap)
    if not chunk_payloads:
        if extraction:
            extraction.status = ContentExtraction.STATUS_FAILED
            extraction.failure_code = "empty_text"
            extraction.error_message = "未解析出可入库文本。"
            extraction.save(update_fields=["status", "failure_code", "error_message", "updated_at"])
        return mark_document_failed(doc, KBDocument.STATUS_FAILED, "未解析出可入库文本。", failure_code="empty_text")

    try:
        with transaction.atomic():
            for index, (content, metadata) in enumerate(chunk_payloads):
                vector = embed_text(content)
                chunk = KBChunk.objects.create(
                    document=doc,
                    kb=kb,
                    chunk_index=index,
                    content=content,
                    metadata=metadata,
                )
                upsert_chunk_indexes(chunk, vector)
            doc.chunk_count = len(chunk_payloads)
            doc.status = KBDocument.STATUS_READY
            doc.error_message = ""
            doc.save(update_fields=["chunk_count", "status", "error_message", "updated_at"])
    except Exception as exc:
        if extraction:
            extraction.status = ContentExtraction.STATUS_FAILED
            extraction.failure_code = "index_failed"
            extraction.error_message = f"索引写入失败：{exc}"[:2000]
            extraction.save(update_fields=["status", "failure_code", "error_message", "updated_at"])
        return mark_document_failed(doc, KBDocument.STATUS_FAILED, f"索引写入失败：{exc}", failure_code="index_failed")

    refresh_kb_doc_count(kb)
    return doc


def ingest_text(kb, source_type, source, title, text, user_file=None, chunk_size=800, overlap=120):
    try:
        entries = entries_from_text(text, title or source, source)
    except DocumentParseError as exc:
        return create_status_document(
            kb,
            source_type,
            source,
            title or source,
            KBDocument.STATUS_FAILED,
            exc,
            user_file=user_file,
            content_hash=short_hash(text),
        )
    return ingest_entries(
        kb,
        source_type,
        source,
        title or source,
        entries,
        user_file=user_file,
        content_hash=short_hash(text),
        chunk_size=chunk_size,
        overlap=overlap,
    )


def ingest_url(kb, url):
    try:
        title, entries = parse_url_entries(url)
    except DocumentParseError as exc:
        return create_status_document(kb, KBDocument.SOURCE_URL, url, url, KBDocument.STATUS_FAILED, exc)
    return ingest_entries(kb, KBDocument.SOURCE_URL, url, title, entries)


def _method_for_parse_error(exc):
    parser_name = exc.parser_name or ""
    if parser_name.startswith("Vision"):
        return ContentExtraction.METHOD_VISION
    if parser_name.startswith(("TextEncodingFallback", "StripRtfExtractor", "SpreadsheetExtractor")):
        return ContentExtraction.METHOD_TEXT_FALLBACK
    if parser_name.startswith("BeautifulSoupFallback"):
        return ContentExtraction.METHOD_HTML_FALLBACK
    if parser_name.startswith(("ArchiveExtractor", "LibreOfficeConverter")):
        return ContentExtraction.METHOD_UNSUPPORTED
    return ContentExtraction.METHOD_LANGCHAIN


def ingest_user_file(kb, user_file):
    if user_file:
        delete_existing_file_documents(kb, user_file)
    try:
        result = parse_user_file_result(user_file)
    except UnsupportedDocumentError as exc:
        profile = exc.profile
        parser_metadata = {
            "profile": profile.as_metadata() if profile else {},
            "parser": exc.parser_name,
            **(getattr(exc, "metadata", {}) or {}),
        }
        extraction = create_content_extraction(
            kb,
            user_file,
            ContentExtraction.STATUS_UNSUPPORTED,
            ContentExtraction.METHOD_UNSUPPORTED,
            failure_code=exc.failure_code,
            error_message=exc,
            metadata=parser_metadata,
        )
        return create_status_document(
            kb,
            KBDocument.SOURCE_FILE,
            user_file.name,
            user_file.name,
            KBDocument.STATUS_UNSUPPORTED,
            exc,
            user_file=user_file,
            content_hash=(user_file.stored_file.content_hash if user_file.stored_file else ""),
            extraction=extraction,
            profile=profile,
            parser_name=exc.parser_name,
            failure_code=exc.failure_code,
            parser_metadata=parser_metadata,
            replace_existing=False,
        )
    except DocumentParseError as exc:
        profile = exc.profile
        parser_metadata = {
            "profile": profile.as_metadata() if profile else {},
            "parser": exc.parser_name,
            **(getattr(exc, "metadata", {}) or {}),
        }
        extraction = create_content_extraction(
            kb,
            user_file,
            ContentExtraction.STATUS_FAILED,
            _method_for_parse_error(exc),
            failure_code=exc.failure_code,
            error_message=exc,
            model_name=settings.VISION_MODEL if (exc.parser_name or "").startswith("Vision") else "",
            metadata=parser_metadata,
        )
        return create_status_document(
            kb,
            KBDocument.SOURCE_FILE,
            user_file.name,
            user_file.name,
            KBDocument.STATUS_FAILED,
            exc,
            user_file=user_file,
            content_hash=(user_file.stored_file.content_hash if user_file.stored_file else ""),
            extraction=extraction,
            profile=profile,
            parser_name=exc.parser_name,
            failure_code=exc.failure_code,
            parser_metadata=parser_metadata,
            replace_existing=False,
        )
    extraction = create_content_extraction(
        kb,
        user_file,
        ContentExtraction.STATUS_READY,
        result.method,
        normalized_text=result.normalized_text,
        raw_output=result.raw_output,
        model_name=result.model_name,
        metadata=result.metadata,
    )
    return ingest_entries(
        kb,
        KBDocument.SOURCE_FILE,
        user_file.name,
        result.title,
        result.entries,
        user_file=user_file,
        content_hash=(user_file.stored_file.content_hash if user_file.stored_file else ""),
        extraction=extraction,
        profile=result.profile,
        parser_name=result.parser_name,
        parser_metadata=result.metadata,
        replace_existing=False,
    )


def _extract_json_queries(text):
    raw = (text or "").strip()
    raw = re.sub(r"^```(?:json)?|```$", "", raw, flags=re.MULTILINE).strip()
    candidates = [raw]
    match = re.search(r"(\{.*\}|\[.*\])", raw, flags=re.DOTALL)
    if match:
        candidates.append(match.group(1))
    for candidate in candidates:
        try:
            parsed = json.loads(candidate)
        except Exception:
            continue
        if isinstance(parsed, dict):
            parsed = parsed.get("queries") or parsed.get("query") or []
        if isinstance(parsed, str):
            parsed = [parsed]
        if isinstance(parsed, list):
            return [compact_text(item) for item in parsed if compact_text(str(item))]
    return []


def _format_chat_history(chat_history):
    if not chat_history:
        return ""
    lines = []
    for message in list(chat_history)[-6:]:
        role = getattr(message, "role", "")
        content = compact_text(getattr(message, "content", ""))
        if content:
            lines.append(f"{role or 'message'}: {content}")
    return "\n".join(lines)


def rewrite_rag_queries(query, chat_history=None):
    query = compact_text(query)
    if not query:
        return []
    if not settings.LLM_API_KEY:
        return [query]
    try:
        client = OpenAI(api_key=settings.LLM_API_KEY, base_url=settings.LLM_BASE_URL)
        history = _format_chat_history(chat_history)
        resp = client.chat.completions.create(
            model=settings.LLM_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "你是 RAG 检索 query 改写器。根据用户问题和少量对话历史，"
                        "生成 1 到 3 个适合知识库语义检索的中文查询。只返回 JSON："
                        "{\"queries\":[\"...\"]}。"
                    ),
                },
                {"role": "user", "content": f"对话历史：\n{history or '无'}\n\n用户问题：{query}"},
            ],
            temperature=0,
        )
        content = resp.choices[0].message.content or ""
        rewritten = _extract_json_queries(content)
    except Exception:
        rewritten = []
    queries = []
    for item in rewritten or [query]:
        item = compact_text(item)
        if item and item not in queries:
            queries.append(item)
        if len(queries) >= 3:
            break
    return queries or [query]


def _candidate_hits(kb, query, top_k, chat_history=None):
    result = _candidate_hits_with_trace(kb, query, top_k=top_k, chat_history=chat_history)
    return result["hits"]


def _candidate_hits_with_trace(kb, query, top_k, chat_history=None):
    queries = rewrite_rag_queries(query, chat_history=chat_history)
    result_limit = max(1, int(top_k))
    candidate_limit = result_limit * 4
    scores = {}
    source_scores = {}
    trace = {
        "original_query": compact_text(query),
        "rewritten_queries": queries,
        "top_k": result_limit,
        "candidate_limit": candidate_limit,
        "vector_candidates": [],
        "fts_candidates": [],
        "fusion_candidates": [],
    }

    for rewritten_query in queries:
        qvec = embed_text(rewritten_query)
        for rank, (chunk_id, distance) in enumerate(vector_candidates(kb, qvec, candidate_limit), 1):
            scores[chunk_id] = scores.get(chunk_id, 0.0) + (1.0 / rank)
            source_scores.setdefault(chunk_id, {})["vector"] = {
                "query": rewritten_query,
                "rank": rank,
                "distance": distance,
            }
            trace["vector_candidates"].append(
                {
                    "query": rewritten_query,
                    "rank": rank,
                    "chunk_id": chunk_id,
                    "distance": distance,
                }
            )
        for rank, (chunk_id, fts_rank) in enumerate(fts_candidates(kb, rewritten_query, candidate_limit), 1):
            scores[chunk_id] = scores.get(chunk_id, 0.0) + (0.5 / rank)
            source_scores.setdefault(chunk_id, {})["fts"] = {
                "query": rewritten_query,
                "rank": rank,
                "rank_score": fts_rank,
            }
            trace["fts_candidates"].append(
                {
                    "query": rewritten_query,
                    "rank": rank,
                    "chunk_id": chunk_id,
                    "rank_score": fts_rank,
                }
            )

    if not scores:
        return {"hits": [], "trace": trace}

    max_candidates = min(100, result_limit * 8)
    ordered_ids = [
        chunk_id
        for chunk_id, _score in sorted(scores.items(), key=lambda item: (-item[1], item[0]))[:max_candidates]
    ]
    chunks = {
        chunk.id: chunk
        for chunk in KBChunk.objects.filter(id__in=ordered_ids).select_related("document", "kb")
    }
    trace["fusion_candidates"] = [
        {
            "rank": rank,
            "chunk_id": chunk_id,
            "score": round(float(scores[chunk_id]), 6),
            "source_scores": source_scores.get(chunk_id, {}),
            "document_title": chunks[chunk_id].document.title if chunk_id in chunks else "",
            "source": chunks[chunk_id].document.source if chunk_id in chunks else "",
        }
        for rank, chunk_id in enumerate(ordered_ids, 1)
    ]
    hits = [
        SearchHit(
            score=scores[chunk_id],
            chunk=chunks[chunk_id],
            query=(source_scores.get(chunk_id, {}).get("vector") or source_scores.get(chunk_id, {}).get("fts") or {}).get(
                "query", query
            ),
            source_scores=source_scores.get(chunk_id, {}),
        )
        for chunk_id in ordered_ids
        if chunk_id in chunks
    ]
    return {"hits": hits, "trace": trace}


def rerank_hits(query, hits, top_k):
    result = rerank_hits_with_trace(query, hits, top_k)
    return result["hits"]


def _trace_hit(hit, rank):
    return {
        "rank": rank,
        "chunk_id": hit.chunk.id,
        "document_id": hit.chunk.document_id,
        "document_title": hit.chunk.document.title,
        "source": hit.chunk.document.source,
        "score": round(float(hit.score), 6),
        "rerank_score": None if hit.rerank_score is None else round(float(hit.rerank_score), 6),
        "source_scores": hit.source_scores,
    }


def rerank_hits_with_trace(query, hits, top_k):
    result_limit = max(1, int(top_k))
    trace = {
        "enabled": bool(getattr(settings, "RERANK_API_KEY", "") or getattr(settings, "EMBEDDING_API_KEY", "")),
        "model": getattr(settings, "RERANK_MODEL", ""),
        "before": [_trace_hit(hit, rank) for rank, hit in enumerate(hits, 1)],
        "after": [],
        "error": "",
    }
    if not hits:
        return {"hits": [], "trace": trace}
    api_key = getattr(settings, "RERANK_API_KEY", "") or getattr(settings, "EMBEDDING_API_KEY", "")
    if not api_key:
        final_hits = hits[:result_limit]
        trace["after"] = [_trace_hit(hit, rank) for rank, hit in enumerate(final_hits, 1)]
        trace["enabled"] = False
        trace["error"] = "missing_api_key"
        return {"hits": final_hits, "trace": trace}

    try:
        payload = {
            "model": settings.RERANK_MODEL,
            "input": {
                "query": {"text": compact_text(query)},
                "documents": [{"text": hit.chunk.content} for hit in hits],
            },
            "parameters": {
                "return_documents": False,
                "top_n": result_limit,
            },
        }
        with httpx.Client(timeout=30) as client:
            response = client.post(
                settings.RERANK_BASE_URL,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json=payload,
            )
            response.raise_for_status()
        results = response.json().get("output", {}).get("results", [])
        reranked = []
        for item in results:
            index = int(item["index"])
            if 0 <= index < len(hits):
                hit = hits[index]
                hit.rerank_score = float(item.get("relevance_score", 0.0))
                reranked.append(hit)
        if reranked:
            final_hits = reranked[:result_limit]
            trace["after"] = [_trace_hit(hit, rank) for rank, hit in enumerate(final_hits, 1)]
            return {"hits": final_hits, "trace": trace}
        trace["error"] = "empty_rerank_results"
    except Exception as exc:
        trace["error"] = str(exc)
    final_hits = hits[:result_limit]
    trace["after"] = [_trace_hit(hit, rank) for rank, hit in enumerate(final_hits, 1)]
    return {"hits": final_hits, "trace": trace}


def search(kb, query, top_k=6, chat_history=None):
    return search_with_trace(kb, query, top_k=top_k, chat_history=chat_history)["hits"]


def search_with_trace(kb, query, top_k=6, chat_history=None):
    candidate_result = _candidate_hits_with_trace(kb, query, top_k=top_k, chat_history=chat_history)
    rerank_result = rerank_hits_with_trace(query, candidate_result["hits"], top_k=top_k)
    trace = candidate_result["trace"]
    trace["rerank"] = rerank_result["trace"]
    trace["final_hits"] = [_trace_hit(hit, rank) for rank, hit in enumerate(rerank_result["hits"], 1)]
    return {"hits": rerank_result["hits"], "trace": trace}


def _hit_parts(hit):
    if isinstance(hit, SearchHit):
        return hit.score, hit.chunk, hit.rerank_score
    score, chunk = hit
    return score, chunk, None


def references_context(hits):
    lines = []
    for index, hit in enumerate(hits, 1):
        score, chunk, rerank_score = _hit_parts(hit)
        score_text = f"{score:.4f}"
        if rerank_score is not None:
            score_text += f", rerank={rerank_score:.4f}"
        lines.append(
            "\n".join(
                [
                    f"[{index}] title: {chunk.document.title}",
                    f"source: {chunk.document.source}",
                    f"chunk_id: {chunk.id}",
                    f"score: {score_text}",
                    "content:",
                    chunk.content,
                ]
            )
        )
    return "\n\n".join(lines)


def build_answer(query, hits, wiki_hits=None):
    if not hits and not wiki_hits:
        return "知识库中暂无足够信息来回答该问题。"
    if wiki_hits:
        from . import wiki_services

        context = wiki_services.combined_references_context(wiki_hits, hits)
    else:
        context = references_context(hits)
    if not settings.LLM_API_KEY:
        return f"未配置 LLM_API_KEY，已返回最相关片段。\n\n问题：{query}\n\n{context[:1800]}"
    try:
        client = OpenAI(api_key=settings.LLM_API_KEY, base_url=settings.LLM_BASE_URL)
        resp = client.chat.completions.create(
            model=settings.LLM_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "你是个人知识库问答助手。优先基于 Wiki references 理解结构化结论，"
                        "再用原文 chunk references 追溯细节。只基于提供的 references 回答；"
                        "如果 references 不足以支持结论，就明确说明资料不足。"
                        "回答要简洁、中文、可追溯。"
                    ),
                },
                {"role": "user", "content": f"用户问题：{query}\n\nreferences:\n{context}"},
            ],
            temperature=0.2,
        )
        return resp.choices[0].message.content or ""
    except Exception as exc:
        return f"模型调用失败，以下是相关片段：\n\n{context[:1800]}\n\n错误：{exc}"


def refs_payload(hits):
    refs = []
    for hit in hits:
        score, chunk, rerank_score = _hit_parts(hit)
        refs.append(
            {
                "type": "chunk",
                "kb_id": chunk.kb.kb_id,
                "document_id": chunk.document_id,
                "chunk_id": chunk.id,
                "title": chunk.document.title,
                "source": chunk.document.source,
                "status": chunk.document.status,
                "score": round(float(score), 4),
                "rerank_score": None if rerank_score is None else round(float(rerank_score), 4),
            }
        )
    return refs


def document_status_label(status, failure_code=""):
    code_labels = {
        "needs_conversion": "需格式转换",
        "conversion_failed": "转换失败",
        "needs_vision": "需视觉解析",
        "vision_not_configured": "需视觉解析",
        "vision_not_supported": "需视觉解析",
        "vision_parse_failed": "视觉解析失败",
        "empty_vision_input": "视觉输入为空",
        "encoding_failed": "编码失败",
        "empty_text": "空文本",
        "spreadsheet_parse_failed": "表格解析失败",
        "sparse_text_fallback_failed": "稀疏文本兜底失败",
        "archive_unsupported": "压缩包不支持",
        "archive_no_supported_files": "压缩包无可入库文件",
        "archive_limit_exceeded": "压缩包超过安全限制",
        "binary_unsupported": "二进制不支持",
        "media_unsupported": "媒体文件不支持",
        "unsupported_format": "不支持",
        "parser_exception": "解析失败",
        "read_failed": "读取失败",
        "index_failed": "索引失败",
    }
    if failure_code in code_labels:
        return code_labels[failure_code]
    return {
        "not_ingested": "未入库",
        KBDocument.STATUS_PROCESSING: "入库中",
        KBDocument.STATUS_READY: "已入库",
        KBDocument.STATUS_FAILED: "解析失败",
        KBDocument.STATUS_UNSUPPORTED: "不支持",
    }.get(status, status or "未入库")


def document_status_map(kb, files):
    file_ids = [file.id for file in files if not file.is_folder]
    latest_docs = {}
    if file_ids:
        docs = KBDocument.objects.filter(kb=kb, user_file_id__in=file_ids).order_by("user_file_id", "-updated_at")
        for doc in docs:
            latest_docs.setdefault(doc.user_file_id, doc)

    status_by_file = {}
    for file in files:
        doc = latest_docs.get(file.id)
        if doc:
            status = doc.status
            status_by_file[file.id] = {
                "status": status,
                "label": document_status_label(status, doc.failure_code),
                "message": doc.error_message,
                "chunk_count": doc.chunk_count,
                "document": doc,
                "failure_code": doc.failure_code,
            }
        else:
            status_by_file[file.id] = {
                "status": "not_ingested",
                "label": document_status_label("not_ingested"),
                "message": "",
                "chunk_count": 0,
                "document": None,
            }
    return status_by_file


def decorate_file_statuses(kb, files):
    items = list(files)
    if not kb:
        return items
    status_by_file = document_status_map(kb, items)
    for item in items:
        info = status_by_file.get(item.id, {})
        item.kb_status = info.get("status", "not_ingested")
        item.kb_status_label = info.get("label", document_status_label("not_ingested"))
        item.kb_status_message = info.get("message", "")
        item.kb_chunk_count = info.get("chunk_count", 0)
        item.kb_failure_code = info.get("failure_code", "")
    return items


def decorate_document_statuses(documents):
    items = list(documents)
    for doc in items:
        doc.kb_status_label = document_status_label(doc.status, doc.failure_code)
        doc.parser_summary = doc.parser_name or (doc.extraction.get_method_display() if doc.extraction_id else "")
    return items
